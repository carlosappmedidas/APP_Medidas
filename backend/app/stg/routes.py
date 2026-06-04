# app/stg/routes.py
# pyright: reportMissingImports=false, reportArgumentType=false
"""
Endpoints REST del módulo STG.

Todos los endpoints requieren autenticación (Bearer JWT) y respetan los
permisos multi-empresa del usuario.
"""
from __future__ import annotations

from datetime import datetime
from io import BytesIO
from typing import Optional

from fastapi import APIRouter, Depends, File, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

from app.core.auth import get_current_user
from app.core.db import get_db
from app.stg import schemas, services
from app.stg.models import (
    StgConcentrador,
    Cups,
)
from app.tenants.models import User


router = APIRouter(prefix="/stg", tags=["stg"])


# ---------------------------------------------------------------------------
# Dashboard
# ---------------------------------------------------------------------------
@router.get("/dashboard/summary", response_model=schemas.DashboardSummary)
def get_dashboard_summary(
    empresa_id: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.get_dashboard_summary(db, user, empresa_id)


# ---------------------------------------------------------------------------
# CUPS
# ---------------------------------------------------------------------------
@router.get("/cups", response_model=schemas.CupsList)
def listar_cups(
    empresa_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    search: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.listar_cups(
        db, user, empresa_id=empresa_id, estado=estado,
        search=search, page=page, page_size=page_size,
    )


@router.get("/cups/{cups_id}", response_model=schemas.CupsRead)
def obtener_cups(
    cups_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    c = db.query(Cups).filter(Cups.id == cups_id).first()
    if c is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "CUPS no encontrado")
    from app.core.permissions import assert_empresa_access
    assert_empresa_access(db, user, c.empresa_id)
    return {
        "id": c.id,
        "empresa_id": c.empresa_id,
        "cups": c.cups,
        "concentrador_id": c.concentrador_id,
        "concentrador_codigo_ct": (
            c.concentrador.codigo_ct if c.concentrador else None
        ),
        "numero_contador": c.numero_contador,
        "fabricante_contador": c.fabricante_contador,
        "modelo_contador": c.modelo_contador,
        "tarifa": c.tarifa,
        "tension_suministro": c.tension_suministro,
        "tipo_punto_medida": c.tipo_punto_medida,
        "direccion": c.direccion,
        "municipio": c.municipio,
        "provincia": c.provincia,
        "cp": c.cp,
        "latitud": c.latitud,
        "longitud": c.longitud,
        "autoconsumo": c.autoconsumo,
        "fecha_alta": c.fecha_alta,
        "fecha_baja": c.fecha_baja,
        "comercializadora_actual": c.comercializadora_actual,
        "ultimo_contacto": c.ultimo_contacto,
        "estado_comunicacion": c.estado_comunicacion,
        "activo": c.activo,
    }


# ---------------------------------------------------------------------------
# Concentradores
# ---------------------------------------------------------------------------
@router.get("/concentradores", response_model=schemas.ConcentradorList)
def listar_concentradores(
    empresa_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.listar_concentradores(
        db, user, empresa_id=empresa_id, estado=estado,
        page=page, page_size=page_size,
    )


@router.get("/concentradores/{concentrador_id}", response_model=schemas.ConcentradorRead)
def obtener_concentrador(
    concentrador_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    c = db.query(StgConcentrador).filter(StgConcentrador.id == concentrador_id).first()
    if c is None:
        raise HTTPException(status.HTTP_404_NOT_FOUND, "Concentrador no encontrado")
    from app.core.permissions import assert_empresa_access
    assert_empresa_access(db, user, c.empresa_id)
    return c


# ---------------------------------------------------------------------------
# Solicitudes
# ---------------------------------------------------------------------------
@router.get("/solicitudes", response_model=schemas.SolicitudFicheroList)
def listar_solicitudes(
    empresa_id: Optional[int] = Query(None),
    estado: Optional[str] = Query(None),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.listar_solicitudes(
        db, user, empresa_id=empresa_id, estado=estado,
        page=page, page_size=page_size,
    )


@router.post("/solicitudes", response_model=schemas.SolicitudFicheroRead, status_code=201)
def crear_solicitud(
    payload: schemas.SolicitudFicheroCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    sol = services.crear_solicitud(db, user, payload.model_dump())
    return {
        "id": sol.id,
        "empresa_id": sol.empresa_id,
        "cups_id": sol.cups_id,
        "cups_codigo": sol.cups.cups if sol.cups else None,
        "concentrador_id": sol.concentrador_id,
        "concentrador_codigo_ct": (
            sol.concentrador.codigo_ct if sol.concentrador else None
        ),
        "tipo_fichero": sol.tipo_fichero,
        "fecha_desde": sol.fecha_desde,
        "fecha_hasta": sol.fecha_hasta,
        "prioridad": sol.prioridad,
        "estado": sol.estado,
        "solicitado_por": sol.solicitado_por,
        "mensaje_error": sol.mensaje_error,
        "fecha_envio": sol.fecha_envio,
        "fecha_recepcion": sol.fecha_recepcion,
        "created_at": sol.created_at,
    }


# ---------------------------------------------------------------------------
# Configuración de conexión por empresa
# ---------------------------------------------------------------------------
@router.get("/conexion", response_model=Optional[schemas.ConexionStgEmpresaRead])
def get_conexion(
    empresa_id: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    conf = services.get_conexion_empresa(db, user, empresa_id)
    return conf


@router.post("/conexion", response_model=schemas.ConexionStgEmpresaRead)
def upsert_conexion(
    payload: schemas.ConexionStgEmpresaCreate,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.upsert_conexion_empresa(db, user, payload.model_dump())


@router.post("/conexion/test", response_model=schemas.ConexionTestResult)
def test_conexion(
    empresa_id: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    return services.probar_conexion(db, user, empresa_id)


# ---------------------------------------------------------------------------
# SFTP — endpoints solo lectura (Paquete 3)
# ---------------------------------------------------------------------------
@router.get("/sftp/listar", response_model=schemas.SftpListadoResponse)
def listar_ficheros_sftp(
    empresa_id: int = Query(...),
    filtro: Optional[str] = Query(None, description="Substring opcional para filtrar nombres"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Lista los ficheros disponibles en la carpeta_recepcion del SFTP de la empresa.

    Solo aplica si la conexión es de tipo 'sftp'. Para otros tipos devuelve
    un listado vacío.
    """
    return services.listar_ficheros_sftp(db, user, empresa_id, filtro_patron=filtro)


# ---------------------------------------------------------------------------
# Descarga real de ficheros (Paquete 5)
# ---------------------------------------------------------------------------
@router.post("/descargar", response_model=schemas.DescargaResponse)
def descargar_ficheros(
    empresa_id: int = Query(...),
    limite: int = Query(5, ge=1, le=1000, description="Máximo a descargar por petición"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Descarga ficheros NUEVOS (que no estén ya en BD) desde el STG remoto
    a `backend/storage/stg/empresa_<id>/<año-mes>/` (o lo que indique
    la env var STG_STORAGE_PATH).

    Filtra duplicados por (empresa_id, nombre_original).
    Extrae metadata del nombre (id_contador, tipo_mensaje, timestamp).

    Aplica a conexiones de tipo 'sftp' o 'ftp'.
    """
    try:
        return services.descargar_ficheros_nuevos(db, user, empresa_id, limite=limite)
    except ValueError as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=str(e))


# ---------------------------------------------------------------------------
# Parseo de ficheros (Paquete 6)
# ---------------------------------------------------------------------------
@router.post("/parsear/{fichero_id}", response_model=schemas.ParseoResponse)
def parsear_fichero(
    fichero_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Parsea un fichero descargado y guarda las medidas en BD.

    - Soporta S24 (vía gisce/primestg).
    - G97 marcado como pendiente (parser propio en futuro paquete).
    - Idempotente: si el fichero ya estaba parsed, se borran las medidas previas
      y se re-parsea.
    """
    try:
        return services.parsear_fichero(db, user, fichero_id)
    except ValueError as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail=str(e))


@router.post("/parsear-pendientes", response_model=schemas.ParseoPendientesResponse)
def parsear_pendientes(
    empresa_id: int = Query(...),
    limite: int = Query(10, ge=1, le=500),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Parsea en bulk hasta `limite` ficheros pendientes (parsed=False) de la empresa.
    Devuelve resumen y detalle por fichero.
    """
    try:
        return services.parsear_pendientes(db, user, empresa_id, limite=limite)
    except ValueError as e:
        from fastapi import HTTPException
        raise HTTPException(status_code=400, detail=str(e))


@router.get("/contadores-detectados", response_model=schemas.ContadoresListResponse)
def listar_contadores_detectados(
    empresa_id: int = Query(...),
    offset: int = Query(0, ge=0, description="Offset de paginación"),
    limit: int = Query(50, ge=1, le=500, description="Tamaño de página (max 500)"),
    concentrador_id: Optional[int] = Query(None, description="Filtrar por concentrador (FK)"),
    estado: Optional[str] = Query(None, description="Filtrar por estado: ok/warning/error/desconocido"),
    fabricante: Optional[str] = Query(None, description="Filtrar por fabricante: CIR/LGZ/SAG/ZIV..."),
    search: Optional[str] = Query(None, description="Texto libre — busca en meter_id o codigo_ct"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Lista los contadores detectados en S24 para una empresa (paginado + filtrado).

    A diferencia de /stg/cups (que lee de stg_cups con código CUPS oficial),
    este endpoint lee de stg_contador y devuelve los contadores físicos
    identificados por su meter_id (CIR..., LGZ..., SAG..., ZIV..., ITE...).

    Filtros (todos opcionales):
      - concentrador_id: FK al concentrador (limitar a un CT concreto)
      - estado:          ok / warning / error / desconocido
      - fabricante:      CIR, LGZ, SAG, ZIV…
      - search:          busca en meter_id O codigo_ct (case-insensitive)

    Los stats globales NO se ven afectados por los filtros — siempre reflejan
    el total de la empresa para mantener el contexto del panorama.
    """
    return services.listar_contadores_detectados(
        db, user, empresa_id,
        offset=offset, limit=limit,
        concentrador_id=concentrador_id,
        estado=estado,
        fabricante=fabricante,
        search=search,
    )


@router.get("/eventos", response_model=schemas.EventosListResponse)
def listar_eventos_humanizados(
    empresa_id: int = Query(..., description="ID de la empresa"),
    meter_id: Optional[str] = Query(None, description="Filtrar por contador (meter_id exacto)"),
    fecha_desde: Optional[datetime] = Query(None, description="Filtrar eventos desde esta fecha (inclusive)"),
    fecha_hasta: Optional[datetime] = Query(None, description="Filtrar eventos hasta esta fecha (inclusive)"),
    limite: int = Query(100, ge=1, le=1000, description="Tamaño de página (max 1000)"),
    offset: int = Query(0, ge=0, description="Offset de paginación"),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """
    Devuelve eventos S09 enriquecidos con descripciones humanas en español.

    Cada evento incluye:
      - grupo + codigo (datos crudos del XML)
      - descripcion_grupo (ej: "Grupo 6 - Alta ocurrencia")
      - descripcion_evento (ej: "Inicio establecimiento de comunicaciones puerto serie")

    También devuelve un `resumen_top` con los 10 tipos de evento más
    frecuentes en el filtro aplicado (útil para dashboards).

    Las descripciones provienen de los diccionarios oficiales de primestg
    (`event_groups`, `meter_events`) y se aplican al renderizar — NO se
    persisten en BD.
    """
    return services.listar_eventos_humanizados(
        db=db,
        user=user,
        empresa_id=empresa_id,
        meter_id=meter_id,
        fecha_desde=fecha_desde,
        fecha_hasta=fecha_hasta,
        limite=limite,
        offset=offset,
    )


# ---------------------------------------------------------------------------
# Import Config — Paquete 8e-2a
# ---------------------------------------------------------------------------
@router.get("/import-config", response_model=schemas.ImportConfigList)
def listar_import_configs(
    empresa_id: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Lista las configs de import para una empresa (max 3 por empresa: excel/gisce_os/sips_cnmc)."""
    return services.listar_import_configs(db, user, empresa_id)


@router.post("/import-config", response_model=schemas.ImportConfigRead)
def upsert_import_config(
    payload: schemas.ImportConfigUpsert,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Crea o actualiza la config de un origen para una empresa. UNIQUE por (empresa_id, origen)."""
    try:
        return services.upsert_import_config(db, user, payload)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.delete("/import-config/{config_id}", status_code=204)
def delete_import_config(
    config_id: int,
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Soft delete (marca activo=False)."""
    services.delete_import_config(db, user, config_id)
    return None


# ---------------------------------------------------------------------------
# Excel Importer endpoints — Paquete 8e-2b
# ---------------------------------------------------------------------------

@router.post("/excel/preview", response_model=schemas.ExcelPreviewResponse)
def preview_excel_endpoint(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Lee cabeceras + 5 filas de muestra de un Excel. NO persiste nada."""
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El fichero debe ser .xlsx o .xls")
    contents = file.file.read()
    try:
        return services.preview_excel(db, user, empresa_id, contents)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


@router.post("/excel/execute", response_model=schemas.ExcelImportResult)
def execute_excel_endpoint(
    empresa_id: int = Query(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Procesa un Excel usando el mapping guardado en stg_import_config (origen=excel).

    Busca cada fila por codigo_ct → actualiza sólo los campos mapeados con valor.
    Si codigo_ct no existe en BD → fila ignorada.
    """
    if not file.filename or not file.filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="El fichero debe ser .xlsx o .xls")
    contents = file.file.read()
    try:
        return services.execute_excel_import(db, user, empresa_id, contents)
    except ValueError as e:
        raise HTTPException(status_code=400, detail=str(e))


# ---------------------------------------------------------------------------
# Excel template export (mapping CIR -> ID_CT) — Paquete 8f-import-export
# ---------------------------------------------------------------------------
@router.get("/concentradores/excel-template")
def get_concentradores_excel_template(
    empresa_id: int = Query(...),
    db: Session = Depends(get_db),
    user: User = Depends(get_current_user),
):
    """Descarga un Excel con los concentradores actuales de la empresa.

    Columnas:
      - codigo_ct  (relleno: el CIR... del concentrador fisico Circutor)
      - id_ct      (VACIO, el cliente lo rellena con el codigo administrativo
                    de la distribuidora, p.ej. '102.CTR.E300000004')
      - nombre_ct  (opcional)
      - nombre     (opcional)

    El cliente rellena 'id_ct' y vuelve a subir el Excel por la pantalla de
    import del 8e-2b. El importer hace UPDATE por (empresa_id, codigo_ct)
    y rellena los huecos.

    Despues, el preview/import GISCE matchea por id_ct == giscedata.cts.name
    y trae nombre/direccion/etc del ERP.
    """
    from app.core.permissions import assert_empresa_access
    assert_empresa_access(db, user, empresa_id)

    concentradores = (
        db.query(StgConcentrador)
        .filter(StgConcentrador.empresa_id == empresa_id)
        .order_by(StgConcentrador.codigo_ct)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    assert ws is not None  # openpyxl siempre crea una hoja activa por defecto
    ws.title = "Concentradores"


    headers = ["codigo_ct", "id_ct", "nombre_ct", "nombre"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="DDDDDD")

    for c in concentradores:
        ws.append([
            c.codigo_ct,
            c.id_ct or "",
            c.nombre_ct or "",
            c.nombre or "",
        ])

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 32
    ws.column_dimensions["C"].width = 40
    ws.column_dimensions["D"].width = 30

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"concentradores_mapping_empresa_{empresa_id}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )
