# app/calendario_ree/routes.py
# pyright: reportMissingImports=false
from __future__ import annotations

import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any, TypedDict, cast

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import FileResponse
from openpyxl import load_workbook
from sqlalchemy import or_
from sqlalchemy.orm import Query as SAQuery, Session

from app.calendario_ree.models import ReeCalendarEvent, ReeCalendarFile
from app.calendario_ree.schemas import (
    ReeCalendarDashboardHitosResponse,
    ReeCalendarFileRead,
    ReeCalendarOperativoItemRead,
    ReeCalendarOperativoResponse,
    ReeCalendarOperativoSeedRequest,
    ReeCalendarWorkbookPreviewResponse,
    ReeCalendarWorkbookSheetRead,
    ReeCalendarWorkbookSheetRowRead,
)
from app.core.auth import get_current_user
from app.core.db import get_db
from app.tenants.models import User

router: APIRouter = APIRouter(prefix="/calendario-ree", tags=["calendario-ree"])
__all__ = ["router"]

UPLOAD_BASE_PATH = Path("data/calendario_ree")
ALLOWED_EXTENSIONS = {".xlsx"}
ALLOWED_MIME_TYPES = {
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "application/octet-stream",
}


class OperativoSeedRow(TypedDict):
    fecha: date
    mes_visual: str
    categoria: str
    evento: str
    mes_afectado: str
    estado: str
    sort_order: int


def _file_any(calendar_file: ReeCalendarFile | None) -> Any:
    return cast(Any, calendar_file)


def _event_any(event: ReeCalendarEvent | None) -> Any:
    return cast(Any, event)


def _safe_filename(filename: str) -> str:
    return Path(filename).name.strip()


def _validate_upload_file(file: UploadFile) -> str:
    if not file.filename:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El fichero debe tener un nombre",
        )

    safe_name = _safe_filename(file.filename)
    ext = Path(safe_name).suffix.lower()

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Formato no permitido. En esta fase solo se admite Excel (.xlsx).",
        )

    if file.content_type and file.content_type not in ALLOWED_MIME_TYPES:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="El fichero debe ser un Excel válido (.xlsx).",
        )

    return safe_name


def _set_calendar_file_values(
    calendar_file: ReeCalendarFile,
    *,
    tenant_id: int,
    anio: int,
    filename: str,
    storage_key: str,
    mime_type: str | None,
    status_value: str,
    is_active: bool,
    uploaded_by: int,
    error_message: str | None,
) -> None:
    calendar_file_any = _file_any(calendar_file)
    calendar_file_any.tenant_id = tenant_id
    calendar_file_any.anio = anio
    calendar_file_any.filename = filename
    calendar_file_any.storage_key = storage_key
    calendar_file_any.mime_type = mime_type
    calendar_file_any.status = status_value
    calendar_file_any.is_active = is_active
    calendar_file_any.uploaded_by = uploaded_by
    calendar_file_any.error_message = error_message


def _set_calendar_event_values(
    event: ReeCalendarEvent,
    *,
    tenant_id: int,
    calendar_file_id: int | None,
    anio: int,
    fecha: date,
    mes_visual: str,
    categoria: str,
    evento: str,
    mes_afectado: str,
    estado: str,
    sort_order: int,
) -> None:
    event_any = _event_any(event)
    event_any.tenant_id = tenant_id
    event_any.calendar_file_id = calendar_file_id
    event_any.anio = anio
    event_any.fecha = fecha
    event_any.mes_visual = mes_visual
    event_any.categoria = categoria
    event_any.evento = evento
    event_any.mes_afectado = mes_afectado
    event_any.estado = estado
    event_any.sort_order = sort_order


def _estado_from_fecha(fecha: date) -> str:
    today = date.today()

    if fecha < today:
        return "cerrado"
    if fecha == today:
        return "hoy"
    if (fecha - today).days <= 15:
        return "proximo"
    return "pendiente"


def _matches_estado_filter(
    estado_value: str,
    estado_filter: str | None,
) -> bool:
    if not estado_filter or estado_filter == "todos":
        return True
    return estado_value == estado_filter


def _is_open_estado(estado_value: str) -> bool:
    return estado_value in {"pendiente", "proximo", "hoy"}


def _to_operativo_schema(item: ReeCalendarEvent) -> ReeCalendarOperativoItemRead:
    item_any = _event_any(item)
    fecha_value = cast(date, item_any.fecha)

    return ReeCalendarOperativoItemRead(
        id=cast(int, item_any.id),
        anio=cast(int, item_any.anio),
        fecha=fecha_value,
        mes_visual=cast(str, item_any.mes_visual),
        categoria=cast(str, item_any.categoria),
        evento=cast(str, item_any.evento),
        mes_afectado=cast(str, item_any.mes_afectado),
        estado=_estado_from_fecha(fecha_value),
        sort_order=cast(int, item_any.sort_order),
    )


def _archive_calendar_file(item: ReeCalendarFile) -> None:
    item_any = _file_any(item)
    item_any.is_active = False
    item_any.status = ReeCalendarFile.STATUS_ARCHIVED


def _activate_calendar_file_model(item: ReeCalendarFile) -> None:
    item_any = _file_any(item)
    item_any.is_active = True
    item_any.status = ReeCalendarFile.STATUS_ACTIVE
    item_any.error_message = None


def _serialize_excel_cell(value: object) -> str:
    if value is None:
        return ""

    if isinstance(value, datetime):
        return value.strftime("%d/%m/%Y %H:%M")
    if isinstance(value, date):
        return value.strftime("%d/%m/%Y")

    return str(value)


def _load_workbook_preview(file_path: Path) -> ReeCalendarWorkbookPreviewResponse:
    workbook = load_workbook(filename=file_path, data_only=True, read_only=True)

    sheets: list[ReeCalendarWorkbookSheetRead] = []

    for worksheet in workbook.worksheets:
        rows: list[ReeCalendarWorkbookSheetRowRead] = []

        max_columns = 0
        for row in worksheet.iter_rows(values_only=True):
            serialized = [_serialize_excel_cell(cell) for cell in row]
            if any(cell != "" for cell in serialized):
                rows.append(ReeCalendarWorkbookSheetRowRead(cells=serialized))
                max_columns = max(max_columns, len(serialized))

        sheets.append(
            ReeCalendarWorkbookSheetRead(
                name=worksheet.title,
                max_columns=max_columns,
                rows=rows,
            )
        )

    workbook.close()
    return ReeCalendarWorkbookPreviewResponse(sheets=sheets)


def _normalize_text(value: str) -> str:
    return " ".join(value.replace("\n", " ").replace("\r", " ").split()).strip()


def _as_date(value: object) -> date | None:
    if value is None:
        return None

    if isinstance(value, datetime):
        return value.date()

    if isinstance(value, date):
        return value

    if isinstance(value, str):
        raw = value.strip()
        if not raw or raw in {"(1)", "(2)"}:
            return None

        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S"):
            try:
                return datetime.strptime(raw, fmt).date()
            except ValueError:
                continue

    return None


def _spanish_month_name(month: int) -> str:
    months = {
        1: "Enero",
        2: "Febrero",
        3: "Marzo",
        4: "Abril",
        5: "Mayo",
        6: "Junio",
        7: "Julio",
        8: "Agosto",
        9: "Septiembre",
        10: "Octubre",
        11: "Noviembre",
        12: "Diciembre",
    }
    return months[month]


def _short_spanish_month_name(month: int) -> str:
    months = {
        1: "Ene",
        2: "Feb",
        3: "Mar",
        4: "Abr",
        5: "May",
        6: "Jun",
        7: "Jul",
        8: "Ago",
        9: "Sep",
        10: "Oct",
        11: "Nov",
        12: "Dic",
    }
    return months[month]


def _month_visual_from_date(value: date) -> str:
    return f"{_spanish_month_name(value.month)} {value.year}"


def _mes_afectado_from_date(value: date) -> str:
    return f"{_spanish_month_name(value.month)} {value.year}"


def _compute_estado(fecha: date) -> str:
    today = date.today()

    if fecha < today:
        return "cerrado"
    if fecha == today:
        return "hoy"
    if (fecha - today).days <= 15:
        return "proximo"
    return "pendiente"


def _build_event_row(
    *,
    fecha: date,
    categoria: str,
    evento: str,
    mes_afectado_date: date,
    sort_order: int,
) -> OperativoSeedRow:
    return {
        "fecha": fecha,
        "mes_visual": _month_visual_from_date(fecha),
        "categoria": categoria,
        "evento": evento,
        "mes_afectado": _mes_afectado_from_date(mes_afectado_date),
        "estado": _compute_estado(fecha),
        "sort_order": sort_order,
    }


def _parse_anexo_ii_sheet(workbook_path: Path) -> list[OperativoSeedRow]:
    workbook = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["ANEXO II (2026)"]
    except KeyError:
        workbook.close()
        return []

    rows = list(worksheet.iter_rows(values_only=True))
    result: list[OperativoSeedRow] = []
    sort_order = 10

    for row in rows[5:18]:
        mes_ref = _as_date(row[8] if len(row) > 8 else None)
        fecha_m1 = _as_date(row[9] if len(row) > 9 else None)
        fecha_m2 = _as_date(row[10] if len(row) > 10 else None)

        if mes_ref and fecha_m1:
            result.append(
                _build_event_row(
                    fecha=fecha_m1,
                    categoria="M+1",
                    evento="Publicación del cierre M+1",
                    mes_afectado_date=mes_ref,
                    sort_order=sort_order,
                )
            )
            sort_order += 10

        if mes_ref and fecha_m2:
            result.append(
                _build_event_row(
                    fecha=fecha_m2,
                    categoria="M+2",
                    evento="Publicación del cierre M+2",
                    mes_afectado_date=mes_ref,
                    sort_order=sort_order,
                )
            )
            sort_order += 10

    header_row = rows[22]
    headers: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if idx >= 2 and cell:
            headers[idx] = _normalize_text(str(cell))

    categoria_by_col: dict[int, str] = {
        3: "Provisional",
        4: "Provisional",
        5: "Provisional",
        6: "Provisional",
        7: "Definitivo",
        8: "Definitivo",
        9: "Definitivo",
        10: "Definitivo",
        11: "Definitivo",
        12: "Definitivo",
        13: "Definitivo",
        14: "Definitivo",
    }

    for row in rows[23:38]:
        mes_ref = _as_date(row[2] if len(row) > 2 else None)
        if not mes_ref:
            continue

        for col_idx in range(3, 15):
            if col_idx >= len(row):
                continue

            fecha_evento = _as_date(row[col_idx])
            if not fecha_evento:
                continue

            header = headers.get(col_idx)
            if not header:
                continue

            categoria = categoria_by_col.get(col_idx, "Operativo")
            result.append(
                _build_event_row(
                    fecha=fecha_evento,
                    categoria=categoria,
                    evento=header,
                    mes_afectado_date=mes_ref,
                    sort_order=sort_order,
                )
            )
            sort_order += 10

    workbook.close()
    return result


def _parse_anexo_iii_sheet(workbook_path: Path) -> list[OperativoSeedRow]:
    workbook = load_workbook(filename=workbook_path, data_only=True, read_only=True)
    try:
        worksheet = workbook["ANEXO III (2026)"]
    except KeyError:
        workbook.close()
        return []

    rows = list(worksheet.iter_rows(values_only=True))
    result: list[OperativoSeedRow] = []
    sort_order = 5000

    header_row = rows[2]
    headers: dict[int, str] = {}
    for idx, cell in enumerate(header_row):
        if cell:
            headers[idx] = _normalize_text(str(cell))

    for row in rows[3:20]:
        mes_ref = _as_date(row[0] if len(row) > 0 else None)
        if not mes_ref:
            continue

        for col_idx in range(1, 12):
            if col_idx >= len(row):
                continue

            fecha_evento = _as_date(row[col_idx])
            if not fecha_evento:
                continue

            header = headers.get(col_idx)
            if not header:
                continue

            result.append(
                _build_event_row(
                    fecha=fecha_evento,
                    categoria="Art. 15",
                    evento=header,
                    mes_afectado_date=mes_ref,
                    sort_order=sort_order,
                )
            )
            sort_order += 10

    workbook.close()
    return result


def _extract_operativo_rows_from_workbook(
    workbook_path: Path,
) -> list[OperativoSeedRow]:
    rows: list[OperativoSeedRow] = []
    rows.extend(_parse_anexo_ii_sheet(workbook_path))
    rows.extend(_parse_anexo_iii_sheet(workbook_path))

    rows.sort(key=lambda item: (item["fecha"], item["sort_order"]))
    return rows


def _replace_operativo_events_for_year(
    db: Session,
    *,
    tenant_id: int,
    anio: int,
    calendar_file_id: int | None,
    workbook_path: Path,
) -> list[ReeCalendarEvent]:
    (
        db.query(ReeCalendarEvent)
        .filter(
            ReeCalendarEvent.tenant_id == tenant_id,
            ReeCalendarEvent.anio == anio,
        )
        .delete(synchronize_session=False)
    )

    rows = _extract_operativo_rows_from_workbook(workbook_path)
    items: list[ReeCalendarEvent] = []

    for row in rows:
        item = ReeCalendarEvent()
        _set_calendar_event_values(
            item,
            tenant_id=tenant_id,
            calendar_file_id=calendar_file_id,
            anio=anio,
            fecha=row["fecha"],
            mes_visual=row["mes_visual"],
            categoria=row["categoria"],
            evento=row["evento"],
            mes_afectado=row["mes_afectado"],
            estado=row["estado"],
            sort_order=row["sort_order"],
        )
        items.append(item)

    db.add_all(items)
    db.commit()

    for item in items:
        db.refresh(item)

    return items


def _apply_operativo_filters(
    query: SAQuery,
    *,
    categoria: str | None,
    estado: str | None,
    search: str | None,
) -> SAQuery:
    if categoria and categoria != "todas":
        query = query.filter(ReeCalendarEvent.categoria == categoria)

    if search:
        search_value = f"%{search.strip()}%"
        query = query.filter(
            or_(
                ReeCalendarEvent.evento.ilike(search_value),
                ReeCalendarEvent.mes_visual.ilike(search_value),
                ReeCalendarEvent.mes_afectado.ilike(search_value),
                ReeCalendarEvent.categoria.ilike(search_value),
            )
        )

    return query


def _build_empty_operativo_response(
    *,
    anio: int | None,
    page_size: int,
) -> ReeCalendarOperativoResponse:
    return ReeCalendarOperativoResponse(
        anio=anio,
        source="db",
        page=1,
        page_size=page_size,
        total=0,
        pages=1,
        total_hitos=0,
        hitos_pendientes=0,
        hitos_cerrados=0,
        categoria_actual=None,
        proximo_hito=None,
        proximos_hitos=[],
        items=[],
    )


def _get_proximos_hitos(
    items: list[ReeCalendarEvent],
    *,
    limit: int = 6,
) -> list[ReeCalendarOperativoItemRead]:
    proximos_models = [
        item
        for item in items
        if _estado_from_fecha(cast(date, _event_any(item).fecha)) != "cerrado"
    ][:limit]

    return [_to_operativo_schema(item) for item in proximos_models]


def _parse_mes_afectado(value: str) -> tuple[int | None, int | None]:
    raw = _normalize_text(value)
    if not raw:
        return None, None

    months = {
        "enero": 1,
        "febrero": 2,
        "marzo": 3,
        "abril": 4,
        "mayo": 5,
        "junio": 6,
        "julio": 7,
        "agosto": 8,
        "septiembre": 9,
        "octubre": 10,
        "noviembre": 11,
        "diciembre": 12,
    }

    parts = raw.split(" ")
    if len(parts) < 2:
        return None, None

    month_value = months.get(parts[0].lower())
    try:
        year_value = int(parts[-1])
    except ValueError:
        return None, None

    return month_value, year_value


def _format_mes_afectado_short(value: str | None) -> str | None:
    if not value:
        return None

    month_value, year_value = _parse_mes_afectado(value)
    if month_value is None or year_value is None:
        return value

    return f"{_short_spanish_month_name(month_value)} {str(year_value)[-2:]}"


def _pick_first_dashboard_item_by_mes_visual(
    items: list[ReeCalendarEvent],
    *,
    mes_visual: str,
    categoria: str | None = None,
    evento_exacto: str | None = None,
    evento_contains: str | None = None,
) -> ReeCalendarEvent | None:
    filtered = [
        item
        for item in items
        if cast(str, _event_any(item).mes_visual) == mes_visual
    ]

    if categoria is not None:
        filtered = [
            item
            for item in filtered
            if cast(str, _event_any(item).categoria) == categoria
        ]

    if evento_exacto is not None:
        filtered = [
            item
            for item in filtered
            if cast(str, _event_any(item).evento) == evento_exacto
        ]

    if evento_contains is not None:
        needle = evento_contains.lower()
        filtered = [
            item
            for item in filtered
            if needle in cast(str, _event_any(item).evento).lower()
        ]

    filtered.sort(
        key=lambda item: (
            cast(date, _event_any(item).fecha),
            cast(int, _event_any(item).sort_order),
            cast(int, _event_any(item).id),
        )
    )

    return filtered[0] if filtered else None


@router.post(
    "/files/upload",
    response_model=ReeCalendarFileRead,
    status_code=status.HTTP_201_CREATED,
)
async def upload_calendar_file(
    anio: int = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    if anio < 2000 or anio > 2100:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="anio debe estar entre 2000 y 2100",
        )

    safe_name = _validate_upload_file(file)
    tenant_id_int = cast(int, current_user.tenant_id)
    uploaded_by_int = cast(int, current_user.id)

    dest_dir = UPLOAD_BASE_PATH / f"tenant_{tenant_id_int}" / str(anio)
    dest_dir.mkdir(parents=True, exist_ok=True)

    dest_path = dest_dir / safe_name
    with dest_path.open("wb") as buffer:
        shutil.copyfileobj(file.file, buffer)

    previous_active = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.tenant_id == tenant_id_int,
            ReeCalendarFile.anio == anio,
            ReeCalendarFile.is_active.is_(True),
        )
        .all()
    )
    for item in previous_active:
        _archive_calendar_file(item)

    calendar_file = ReeCalendarFile()
    _set_calendar_file_values(
        calendar_file,
        tenant_id=tenant_id_int,
        anio=anio,
        filename=safe_name,
        storage_key=str(dest_path),
        mime_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        status_value=ReeCalendarFile.STATUS_ACTIVE,
        is_active=True,
        uploaded_by=uploaded_by_int,
        error_message=None,
    )

    db.add(calendar_file)
    db.commit()
    db.refresh(calendar_file)

    _replace_operativo_events_for_year(
        db,
        tenant_id=tenant_id_int,
        anio=anio,
        calendar_file_id=cast(int, _file_any(calendar_file).id),
        workbook_path=dest_path,
    )

    db.refresh(calendar_file)
    return calendar_file


@router.get("/files", response_model=list[ReeCalendarFileRead])
def list_calendar_files(
    anio: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)

    query = db.query(ReeCalendarFile).filter(
        ReeCalendarFile.tenant_id == tenant_id_int,
    )

    if anio is not None:
        query = query.filter(ReeCalendarFile.anio == anio)

    return query.order_by(
        ReeCalendarFile.anio.desc(),
        ReeCalendarFile.created_at.desc(),
        ReeCalendarFile.id.desc(),
    ).all()


@router.get(
    "/files/{file_id}/preview",
    response_model=ReeCalendarWorkbookPreviewResponse,
)
def get_calendar_file_preview(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)

    calendar_file = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.id == file_id,
            ReeCalendarFile.tenant_id == tenant_id_int,
        )
        .first()
    )

    if not calendar_file:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calendario REE no encontrado",
        )

    storage_key = cast(str | None, _file_any(calendar_file).storage_key)
    if not storage_key:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="El calendario no tiene preview disponible",
        )

    file_path = Path(storage_key)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encuentra el fichero del calendario",
        )

    return _load_workbook_preview(file_path)


@router.get("/files/{file_id}/download")
def download_calendar_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)

    calendar_file = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.id == file_id,
            ReeCalendarFile.tenant_id == tenant_id_int,
        )
        .first()
    )

    if not calendar_file:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calendario REE no encontrado",
        )

    storage_key = cast(str | None, _file_any(calendar_file).storage_key)
    if not storage_key:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="El calendario no tiene fichero descargable",
        )

    file_path = Path(storage_key)
    if not file_path.exists() or not file_path.is_file():
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="No se encuentra el fichero del calendario",
        )

    filename = cast(str, _file_any(calendar_file).filename)

    return FileResponse(
        path=file_path,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        filename=filename,
    )


@router.post("/operativo/seed", response_model=ReeCalendarOperativoResponse)
def seed_calendar_operativo(
    payload: ReeCalendarOperativoSeedRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)
    anio = payload.anio

    calendar_file = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.tenant_id == tenant_id_int,
            ReeCalendarFile.anio == anio,
            ReeCalendarFile.is_active.is_(True),
        )
        .order_by(
            ReeCalendarFile.created_at.desc(),
            ReeCalendarFile.id.desc(),
        )
        .first()
    )

    if calendar_file is None:
        return _build_empty_operativo_response(anio=anio, page_size=25)

    storage_key = cast(str | None, _file_any(calendar_file).storage_key)
    if not storage_key:
        return _build_empty_operativo_response(anio=anio, page_size=25)

    items = _replace_operativo_events_for_year(
        db,
        tenant_id=tenant_id_int,
        anio=anio,
        calendar_file_id=cast(int, _file_any(calendar_file).id),
        workbook_path=Path(storage_key),
    )

    sorted_items = sorted(
        items,
        key=lambda item: (
            cast(date, _event_any(item).fecha),
            cast(int, _event_any(item).sort_order),
            cast(int, _event_any(item).id),
        ),
    )

    hitos_pendientes = len(
        [
            item
            for item in sorted_items
            if _is_open_estado(_estado_from_fecha(cast(date, _event_any(item).fecha)))
        ]
    )
    hitos_cerrados = len(
        [
            item
            for item in sorted_items
            if _estado_from_fecha(cast(date, _event_any(item).fecha)) == "cerrado"
        ]
    )

    proximos_hitos_models = [
        item
        for item in sorted_items
        if _is_open_estado(_estado_from_fecha(cast(date, _event_any(item).fecha)))
    ][:6]

    proximo_hito_model = proximos_hitos_models[0] if proximos_hitos_models else None
    proximo_hito = (
        _to_operativo_schema(proximo_hito_model) if proximo_hito_model is not None else None
    )
    categoria_actual = (
        cast(str, _event_any(proximo_hito_model).categoria)
        if proximo_hito_model is not None
        else None
    )

    total_hitos = len(sorted_items)

    return ReeCalendarOperativoResponse(
        anio=anio,
        source="db",
        page=1,
        page_size=25,
        total=total_hitos,
        pages=1 if total_hitos == 0 else (total_hitos + 24) // 25,
        total_hitos=total_hitos,
        hitos_pendientes=hitos_pendientes,
        hitos_cerrados=hitos_cerrados,
        categoria_actual=categoria_actual,
        proximo_hito=proximo_hito,
        proximos_hitos=[_to_operativo_schema(item) for item in proximos_hitos_models],
        items=[_to_operativo_schema(item) for item in sorted_items[:25]],
    )


@router.get("/operativo", response_model=ReeCalendarOperativoResponse)
def get_calendar_operativo(
    anio: int | None = None,
    categoria: str | None = None,
    estado: str | None = None,
    search: str | None = None,
    page: int = Query(default=1, ge=1),
    page_size: int = Query(default=25, ge=1, le=100),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)

    selected_anio = anio
    if selected_anio is None:
        active_calendar = (
            db.query(ReeCalendarFile)
            .filter(
                ReeCalendarFile.tenant_id == tenant_id_int,
                ReeCalendarFile.is_active.is_(True),
            )
            .order_by(
                ReeCalendarFile.anio.desc(),
                ReeCalendarFile.created_at.desc(),
                ReeCalendarFile.id.desc(),
            )
            .first()
        )

        if active_calendar is not None:
            selected_anio = cast(int, _file_any(active_calendar).anio)

    if selected_anio is None:
        return _build_empty_operativo_response(anio=None, page_size=page_size)

    base_query = (
        db.query(ReeCalendarEvent)
        .filter(
            ReeCalendarEvent.tenant_id == tenant_id_int,
            ReeCalendarEvent.anio == selected_anio,
        )
    )

    filtered_query = _apply_operativo_filters(
        base_query,
        categoria=categoria,
        estado=estado,
        search=search,
    )

    db_items = (
        filtered_query.order_by(
            ReeCalendarEvent.fecha.asc(),
            ReeCalendarEvent.sort_order.asc(),
            ReeCalendarEvent.id.asc(),
        )
        .all()
    )

    estado_filtered_items = [
        item
        for item in db_items
        if _matches_estado_filter(
            _estado_from_fecha(cast(date, _event_any(item).fecha)),
            estado,
        )
    ]

    total = len(estado_filtered_items)
    pages = max(1, (total + page_size - 1) // page_size)
    safe_page = page if page <= pages else pages
    offset = (safe_page - 1) * page_size
    paged_items = estado_filtered_items[offset : offset + page_size]

    hitos_pendientes = len(
        [
            item
            for item in estado_filtered_items
            if _is_open_estado(_estado_from_fecha(cast(date, _event_any(item).fecha)))
        ]
    )
    hitos_cerrados = len(
        [
            item
            for item in estado_filtered_items
            if _estado_from_fecha(cast(date, _event_any(item).fecha)) == "cerrado"
        ]
    )

    proximo_hito_model = next(
        (
            item
            for item in estado_filtered_items
            if _is_open_estado(_estado_from_fecha(cast(date, _event_any(item).fecha)))
        ),
        None,
    )

    proximo_hito = (
        _to_operativo_schema(proximo_hito_model) if proximo_hito_model is not None else None
    )

    categoria_actual = (
        cast(str, _event_any(proximo_hito_model).categoria)
        if proximo_hito_model is not None
        else None
    )

    proximos_hitos = _get_proximos_hitos(estado_filtered_items, limit=5)

    return ReeCalendarOperativoResponse(
        anio=selected_anio,
        source="db",
        page=safe_page,
        page_size=page_size,
        total=total,
        pages=pages,
        total_hitos=total,
        hitos_pendientes=hitos_pendientes,
        hitos_cerrados=hitos_cerrados,
        categoria_actual=categoria_actual,
        proximo_hito=proximo_hito,
        proximos_hitos=proximos_hitos,
        items=[_to_operativo_schema(item) for item in paged_items],
    )


@router.get("/dashboard-hitos", response_model=ReeCalendarDashboardHitosResponse)
def get_dashboard_hitos(
    anio: int | None = None,
    mes: int | None = Query(default=None, ge=1, le=12),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)
    today = date.today()

    target_anio = anio if anio is not None else today.year
    target_mes = mes if mes is not None else today.month
    target_mes_label = f"{_spanish_month_name(target_mes)} {target_anio}"

    all_items = (
        db.query(ReeCalendarEvent)
        .filter(
            ReeCalendarEvent.tenant_id == tenant_id_int,
            ReeCalendarEvent.anio == target_anio,
        )
        .order_by(
            ReeCalendarEvent.fecha.asc(),
            ReeCalendarEvent.sort_order.asc(),
            ReeCalendarEvent.id.asc(),
        )
        .all()
    )

    item_m2 = _pick_first_dashboard_item_by_mes_visual(
        all_items,
        mes_visual=target_mes_label,
        categoria="M+2",
        evento_contains="cierre m+2",
    )

    item_m7 = _pick_first_dashboard_item_by_mes_visual(
        all_items,
        mes_visual=target_mes_label,
        evento_contains="cierre provisional",
    )

    item_limite_obj = _pick_first_dashboard_item_by_mes_visual(
        all_items,
        mes_visual=target_mes_label,
        evento_contains="FIN RESOLUCIÓN OBJECIONES",
    )

    item_m11 = _pick_first_dashboard_item_by_mes_visual(
        all_items,
        mes_visual=target_mes_label,
        evento_contains="cierre definitivo",
    )

    item_art15 = _pick_first_dashboard_item_by_mes_visual(
        all_items,
        mes_visual=target_mes_label,
        categoria="Art. 15",
        evento_contains="publicación del operador del sistema",
    )

    return ReeCalendarDashboardHitosResponse(
        anio=target_anio,
        mes=target_mes,
        mes_label=target_mes_label,
        fecha_publicacion_m2=cast(date | None, _event_any(item_m2).fecha) if item_m2 else None,
        mes_afectado_publicacion_m2=_format_mes_afectado_short(
            cast(str | None, _event_any(item_m2).mes_afectado) if item_m2 else None
        ),
        fecha_publicacion_m7=cast(date | None, _event_any(item_m7).fecha) if item_m7 else None,
        mes_afectado_publicacion_m7=_format_mes_afectado_short(
            cast(str | None, _event_any(item_m7).mes_afectado) if item_m7 else None
        ),
        fecha_limite_respuesta_objeciones=cast(date | None, _event_any(item_limite_obj).fecha) if item_limite_obj else None,
        mes_afectado_limite_respuesta_objeciones=_format_mes_afectado_short(
            cast(str | None, _event_any(item_limite_obj).mes_afectado) if item_limite_obj else None
        ),
        fecha_publicacion_m11=cast(date | None, _event_any(item_m11).fecha) if item_m11 else None,
        mes_afectado_publicacion_m11=_format_mes_afectado_short(
            cast(str | None, _event_any(item_m11).mes_afectado) if item_m11 else None
        ),
        fecha_publicacion_art15=cast(date | None, _event_any(item_art15).fecha) if item_art15 else None,
        mes_afectado_publicacion_art15=_format_mes_afectado_short(
            cast(str | None, _event_any(item_art15).mes_afectado) if item_art15 else None
        ),
    )


@router.post(
    "/files/{file_id}/activate",
    response_model=ReeCalendarFileRead,
    status_code=status.HTTP_200_OK,
)
def activate_calendar_file(
    file_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user),
):
    tenant_id_int = cast(int, current_user.tenant_id)

    calendar_file = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.id == file_id,
            ReeCalendarFile.tenant_id == tenant_id_int,
        )
        .first()
    )

    if not calendar_file:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="Calendario REE no encontrado",
        )

    calendar_file_anio = cast(int, _file_any(calendar_file).anio)

    current_active = (
        db.query(ReeCalendarFile)
        .filter(
            ReeCalendarFile.tenant_id == tenant_id_int,
            ReeCalendarFile.anio == calendar_file_anio,
            ReeCalendarFile.is_active.is_(True),
            ReeCalendarFile.id != file_id,
        )
        .all()
    )

    for item in current_active:
        _archive_calendar_file(item)

    _activate_calendar_file_model(calendar_file)
    db.commit()
    db.refresh(calendar_file)

    storage_key = cast(str | None, _file_any(calendar_file).storage_key)
    if storage_key:
        _replace_operativo_events_for_year(
            db,
            tenant_id=tenant_id_int,
            anio=calendar_file_anio,
            calendar_file_id=cast(int, _file_any(calendar_file).id),
            workbook_path=Path(storage_key),
        )

    return calendar_file