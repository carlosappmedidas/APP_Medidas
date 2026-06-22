# app/erp/migraciones/informe.py
"""
Genera el Excel de RESULTADO de una migración (E-12d).

Recibe el dict del resultado (el mismo que ResultadoImport.as_dict() y que el
frontend ya tiene tras importar) y produce un .xlsx con dos hojas:
  - "Resumen": entidad, total, creadas, omitidas, fallidas (+ errores de fichero).
  - "Errores": una fila por error (entidad, fila, columna, valor, motivo).

No reimporta nada: solo formatea el resultado que se mostró en pantalla.
Acepta un único resultado o una lista de resultados (varias entidades).
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

_FONT = "Calibri"
_HDR_FILL = "1F4E78"      # azul cabecera
_HDR_FONT = "FFFFFF"
_OK_FILL = "C6E0B4"       # verde (creadas)
_WARN_FILL = "FFE699"     # amarillo (omitidas)
_ERR_FILL = "F8CBAD"      # rojo suave (fallidas)


def _cab(ws, fila, textos, anchos):
    for i, (txt, ancho) in enumerate(zip(textos, anchos), start=1):
        c = ws.cell(row=fila, column=i, value=txt)
        c.font = Font(name=_FONT, bold=True, size=10, color=_HDR_FONT)
        c.fill = PatternFill("solid", fgColor=_HDR_FILL)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[c.column_letter].width = ancho


def _norm(resultado):
    """Acepta un dict de resultado o una lista; devuelve siempre lista."""
    if isinstance(resultado, dict):
        return [resultado]
    return list(resultado or [])


def generar_informe(resultado) -> bytes:
    """Devuelve el .xlsx (bytes) del informe de migración."""
    resultados = _norm(resultado)

    wb = Workbook()

    # --- Hoja Resumen ---
    ws = wb.active
    ws.title = "Resumen"
    t = ws.cell(row=1, column=1, value="Informe de migración")
    t.font = Font(name=_FONT, bold=True, size=14)
    ws.merge_cells("A1:E1")

    _cab(ws, 3, ["Entidad", "Total", "Creadas", "Omitidas", "Fallidas"], [26, 10, 12, 12, 12])
    r = 4
    tot = {"total": 0, "creadas": 0, "omitidas": 0, "fallidas": 0}
    for res in resultados:
        ws.cell(row=r, column=1, value=res.get("entidad", "")).font = Font(name=_FONT, size=10)
        for col, clave, fill in (
            (2, "total", None), (3, "creadas", _OK_FILL),
            (4, "omitidas", _WARN_FILL), (5, "fallidas", _ERR_FILL),
        ):
            val = int(res.get(clave, 0) or 0)
            tot[clave] = tot.get(clave, 0) + val
            c = ws.cell(row=r, column=col, value=val)
            c.font = Font(name=_FONT, size=10)
            c.alignment = Alignment(horizontal="center")
            if fill and val:
                c.fill = PatternFill("solid", fgColor=fill)
        r += 1

    # fila de totales
    ct = ws.cell(row=r, column=1, value="TOTAL")
    ct.font = Font(name=_FONT, bold=True, size=10)
    for col, clave in ((2, "total"), (3, "creadas"), (4, "omitidas"), (5, "fallidas")):
        c = ws.cell(row=r, column=col, value=tot.get(clave, 0))
        c.font = Font(name=_FONT, bold=True, size=10)
        c.alignment = Alignment(horizontal="center")

    # errores de fichero (si los hay), debajo
    rr = r + 2
    fich = [(res.get("entidad", ""), m) for res in resultados for m in (res.get("errores_fichero") or [])]
    if fich:
        h = ws.cell(row=rr, column=1, value="Errores de fichero")
        h.font = Font(name=_FONT, bold=True, size=11)
        rr += 1
        for ent, msg in fich:
            ws.cell(row=rr, column=1, value=ent).font = Font(name=_FONT, size=10)
            mc = ws.cell(row=rr, column=2, value=msg)
            mc.font = Font(name=_FONT, size=10)
            mc.alignment = Alignment(wrap_text=True)
            rr += 1

    # --- Hoja Errores ---
    we = wb.create_sheet("Errores")
    _cab(we, 1, ["Entidad", "Fila", "Columna", "Valor", "Motivo"], [22, 8, 24, 28, 70])
    er = 2
    for res in resultados:
        ent = res.get("entidad", "")
        for e in (res.get("errores") or []):
            ws_vals = [ent, e.get("fila"), e.get("columna"), e.get("valor"), e.get("motivo")]
            for col, val in enumerate(ws_vals, start=1):
                c = we.cell(row=er, column=col, value="" if val is None else str(val))
                c.font = Font(name=_FONT, size=10)
                c.alignment = Alignment(vertical="top", wrap_text=(col == 5))
            er += 1
    if er == 2:
        c = we.cell(row=2, column=1, value="Sin errores 🎉")
        c.font = Font(name=_FONT, size=10)
    we.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
