# app/erp/migraciones/lectura.py
"""
Lectura de un Excel de migración (E-12c).

Abre el .xlsx subido, localiza la hoja de datos (la que no es "Instrucciones"),
lee la fila 1 como cabeceras y las filas desde la 2 como datos, devolviendo cada
fila como dict {columna: valor} junto con su nº de fila ORIGINAL del Excel
(para que el informe de errores señale exactamente dónde corregir).

No sabe nada de empresa ni de negocio: solo parsea. La validación y la resolución
de claves naturales las hace el importer.
"""
from __future__ import annotations

import io
from dataclasses import dataclass, field

from openpyxl import load_workbook

HOJA_INSTRUCCIONES = "Instrucciones"


@dataclass
class FilaLeida:
    fila_excel: int                 # nº de fila real en el Excel (1-indexado; datos empiezan en 2)
    valores: dict[str, object]      # {cabecera: valor} ya saneado


@dataclass
class LecturaResultado:
    hoja: str
    cabeceras: list[str]
    filas: list[FilaLeida] = field(default_factory=list)
    errores: list[str] = field(default_factory=list)   # errores a nivel de fichero (no de fila)


def _limpiar(valor):
    """Normaliza el valor de una celda: None/'' -> None; strings -> strip."""
    if valor is None:
        return None
    if isinstance(valor, str):
        v = valor.strip()
        return v if v != "" else None
    return valor


def leer_excel(contenido: bytes) -> LecturaResultado:
    """Lee el .xlsx (bytes) y devuelve cabeceras + filas de datos de la hoja de datos."""
    try:
        wb = load_workbook(io.BytesIO(contenido), read_only=True, data_only=True)
    except Exception as e:  # fichero no es un xlsx válido
        return LecturaResultado(hoja="", cabeceras=[], errores=[f"No se pudo abrir el fichero Excel: {e}"])

    # Hoja de datos = la primera que no sea "Instrucciones".
    hojas_datos = [n for n in wb.sheetnames if n != HOJA_INSTRUCCIONES]
    if not hojas_datos:
        wb.close()
        return LecturaResultado(hoja="", cabeceras=[], errores=["El Excel no tiene hoja de datos (solo Instrucciones)."])
    nombre_hoja = hojas_datos[0]
    ws = wb[nombre_hoja]

    filas_iter = ws.iter_rows(values_only=True)
    try:
        cabecera_row = next(filas_iter)
    except StopIteration:
        wb.close()
        return LecturaResultado(hoja=nombre_hoja, cabeceras=[], errores=["La hoja de datos está vacía."])

    cabeceras = [str(c).strip() if c is not None else "" for c in cabecera_row]
    while cabeceras and cabeceras[-1] == "":
        cabeceras.pop()

    res = LecturaResultado(hoja=nombre_hoja, cabeceras=cabeceras)
    if not cabeceras:
        res.errores.append("La fila de cabeceras está vacía.")
        wb.close()
        return res

    nfila = 1  # cabeceras = fila 1
    for row in filas_iter:
        nfila += 1
        valores = {}
        hay_algo = False
        for i, cab in enumerate(cabeceras):
            if not cab:
                continue
            v = _limpiar(row[i]) if i < len(row) else None
            valores[cab] = v
            if v is not None:
                hay_algo = True
        if not hay_algo:
            continue  # fila completamente vacía: se ignora (no es error)
        res.filas.append(FilaLeida(fila_excel=nfila, valores=valores))

    wb.close()
    return res
