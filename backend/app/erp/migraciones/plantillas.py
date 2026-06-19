# app/erp/migraciones/plantillas.py
"""
Generación de plantillas Excel para la migración de empresas (E-12a).

Cada plantilla = un .xlsx con dos hojas:
  - "Instrucciones": qué es, orden de carga, leyenda de colores y ayuda campo a campo.
  - hoja de datos de la entidad: cabeceras coloreadas + comentarios + fila 2 de ejemplo.

Los colores/comentarios son solo visuales: al importar se leen únicamente los
valores (cabecera para mapear + datos desde la fila 2). No afectan a la carga.
"""
from __future__ import annotations

import io

from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill

_FONT = "Calibri"

_FILL = {
    "clave":    "C6E0B4",
    "enlace":   "BDD7EE",
    "derivado": "FFE699",
    "normal":   "D9D9D9",
}
_COMMENT = {
    "clave":    "CLAVE: identifica esta fila de forma única en la empresa. Obligatoria, no repetir.",
    "enlace":   "ENLACE por clave natural: debe existir ya (cargado en su plantilla previa). El sistema lo resuelve solo.",
    "derivado": "AUTOMÁTICO: lo calcula el sistema al guardar. NO rellenar.",
}

PLANTILLAS: dict[str, dict] = {
    "titulares": {
        "hoja": "Titulares",
        "titulo": "Plantilla de migración — TITULARES",
        "que_es": "Carga de una vez todos los titulares (clientes) de la empresa al migrarla al ERP.",
        "orden": "Los titulares se cargan PRIMERO (junto con las comercializadoras de empresa). Suministros y contratos los necesitan.",
        "columnas": [
            ("tipo_persona", "fisica", "normal", "fisica / juridica."),
            ("tipo_identificador", "NI", "normal", "CI(CIF) / DN(DNI/NIF) / NI(NIE) / OT(otro) / NE(extranjero sin NIE)."),
            ("identificador", "12345678Z", "clave", "NIF/CIF/NIE. Clave única del titular en la empresa. Obligatorio."),
            ("nombre_de_pila", "María", "normal", "Solo persona física."),
            ("primer_apellido", "García", "normal", "Solo persona física."),
            ("segundo_apellido", "López", "normal", "Solo persona física."),
            ("razon_social", "", "normal", "Solo persona jurídica."),
            ("nombre", "(automático)", "derivado", "Se compone solo (pila+apellidos o razón social). NO rellenar."),
            ("dir_tipo_via", "CL", "normal", "Código CNMC de tipo de vía (CL, AV, PZ...)."),
            ("dir_via", "Mayor", "normal", "Nombre de la vía."),
            ("dir_numero", "12", "normal", "Número (o SN si sin número)."),
            ("dir_duplicador", "", "normal", "Bis, duplicado..."),
            ("dir_escalera", "", "normal", ""),
            ("dir_piso", "2", "normal", ""),
            ("dir_puerta", "B", "normal", ""),
            ("dir_tipo_aclarador", "", "normal", "Código CNMC de aclarador de finca."),
            ("dir_aclarador", "", "normal", ""),
            ("dir_cp", "28013", "normal", "Código postal (5 dígitos)."),
            ("dir_municipio", "Madrid", "normal", ""),
            ("dir_provincia", "Madrid", "normal", ""),
            ("dir_pais", "España", "normal", ""),
            ("persona_contacto", "Juan Pérez", "normal", ""),
            ("telefono", "910000000", "normal", ""),
            ("movil", "600000000", "normal", ""),
            ("email", "maria@ejemplo.es", "normal", "Debe ser un email válido."),
            ("codigo_interno", "", "normal", "Código propio opcional."),
            ("notas", "", "normal", ""),
            ("activo", "TRUE", "normal", "TRUE / FALSE."),
        ],
    },
    "comercializadoras_empresa": {
        "hoja": "ComercializadorasEmpresa",
        "titulo": "Plantilla de migración — COMERCIALIZADORAS DE LA EMPRESA",
        "que_es": "Da de alta en esta empresa las comercializadoras (con su forma de pago) que ya existen en el catálogo global. NO crea comercializadoras nuevas en el catálogo global (eso lo hace el administrador).",
        "orden": "Se cargan al principio (junto con titulares). Los contratos las necesitan.",
        "columnas": [
            ("comercializadora_codigo_ree", "0123", "enlace", "Código REE de la comercializadora (debe existir en el catálogo global). Clave de enlace."),
            ("direccion", "Calle Ejemplo 1, Madrid", "normal", "Dirección de la comercializadora para esta empresa."),
            ("tipo_pago", "Transferencia 30 días", "normal", "Forma de pago acordada (texto libre)."),
            ("fecha_alta_erp", "2024-01-01", "normal", "AAAA-MM-DD."),
            ("fecha_baja_erp", "", "normal", "AAAA-MM-DD (vacío si activa)."),
            ("activo", "TRUE", "normal", "TRUE / FALSE."),
        ],
    },
    "suministros": {
        "hoja": "Suministros",
        "titulo": "Plantilla de migración — SUMINISTROS",
        "que_es": "Carga de una vez todos los puntos de suministro (CUPS) de la empresa.",
        "orden": "Se cargan DESPUÉS de titulares. Los contratos los necesitan.",
        "columnas": [
            ("cups", "ES0031000000000123AB", "clave", "CUPS (ES + 16 dígitos + 2 letras). Clave única del suministro en la empresa."),
            ("distribuidora", "", "normal", "Normalmente la propia empresa."),
            ("acometida", "", "normal", ""),
            ("dir_tipo_via", "CL", "normal", "Código CNMC de tipo de vía. Obligatorio."),
            ("dir_via", "Mayor", "normal", "Obligatorio."),
            ("dir_numero", "12", "normal", "Obligatorio (o SN)."),
            ("dir_duplicador", "", "normal", ""),
            ("dir_escalera", "", "normal", ""),
            ("dir_piso", "", "normal", ""),
            ("dir_puerta", "", "normal", ""),
            ("dir_tipo_aclarador", "", "normal", ""),
            ("dir_aclarador", "", "normal", ""),
            ("dir_cp", "28013", "normal", "Obligatorio."),
            ("dir_municipio", "Madrid", "normal", "Obligatorio."),
            ("dir_poblacion", "Madrid", "normal", "Obligatorio."),
            ("dir_provincia", "Madrid", "normal", "Obligatorio."),
            ("dir_pais", "España", "normal", "Obligatorio."),
            ("municipio_codigo_ine", "28079", "normal", "Código INE del municipio. Obligatorio."),
            ("poligono", "", "normal", "Catastro."),
            ("parcela", "", "normal", "Catastro."),
            ("ref_catastral", "", "normal", "Referencia catastral (20 caracteres)."),
            ("utm_x", "", "normal", "Coordenada UTM X."),
            ("utm_y", "", "normal", "Coordenada UTM Y."),
            ("utm_huso", "", "normal", "Huso UTM (28-31 en España)."),
            ("utm_banda", "", "normal", "Banda UTM (una letra)."),
            ("latitud", "", "normal", "-90 a 90."),
            ("longitud", "", "normal", "-180 a 180."),
            ("zona", "", "normal", ""),
            ("orden", "", "normal", ""),
            ("centro_transformador", "", "normal", ""),
            ("linea", "", "normal", ""),
            ("pot_max_admisible_cie_kw", "5.5", "normal", "kW. Obligatorio."),
            ("potencia_adscrita_kw", "5.5", "normal", "kW. Obligatorio."),
            ("potencia_adscrita_bloqueada", "FALSE", "normal", "TRUE / FALSE."),
            ("fecha_vigencia_adscrita", "", "normal", "AAAA-MM-DD."),
            ("potencia_convenio_kw", "", "normal", "kW."),
            ("criterio_regulatorio", "", "normal", ""),
            ("fecha_alta", "2024-01-15", "normal", "AAAA-MM-DD."),
            ("fecha_baja", "", "normal", "AAAA-MM-DD."),
            ("notas", "", "normal", ""),
            ("activo", "TRUE", "normal", "TRUE / FALSE."),
        ],
    },
    "contratos": {
        "hoja": "Contratos",
        "titulo": "Plantilla de migración — CONTRATOS",
        "que_es": "Carga de una vez todos los contratos de la empresa.",
        "orden": "Se cargan AL FINAL: necesitan que titular, suministro y comercializadora de empresa ya existan.",
        "columnas": [
            ("numero_contrato", "CTR-2024-000123", "clave", "Clave única del contrato en la empresa. Obligatorio."),
            ("titular_identificador", "12345678Z", "enlace", "NIF/CIF del titular (debe existir ya)."),
            ("pagador_identificador", "", "enlace", "NIF/CIF del pagador si difiere del titular. Opcional."),
            ("suministro_cups", "ES0031000000000123AB", "enlace", "CUPS del suministro (debe existir ya)."),
            ("comercializadora_codigo_ree", "0123", "enlace", "Código REE; debe estar dada de alta en la empresa."),
            ("tarifa_codigo", "2.0TD", "enlace", "Código de tarifa de acceso (2.0TD, 3.0TD, 6.1TD...)."),
            ("tipo_punto_medida", "(automático)", "derivado", "Lo calcula el sistema según la potencia. NO rellenar."),
            ("tipo_contrato_atr", "anual", "normal", "anual / eventual / temporada / obras."),
            ("estado", "activo", "normal", "borrador / activo / baja."),
            ("fecha_alta", "2024-01-15", "normal", "AAAA-MM-DD."),
            ("fecha_activacion_prevista", "", "normal", "AAAA-MM-DD."),
            ("fecha_firma", "", "normal", "AAAA-MM-DD."),
            ("fecha_baja", "", "normal", "AAAA-MM-DD."),
            ("fecha_finalizacion", "", "normal", "AAAA-MM-DD."),
            ("renovacion_automatica", "FALSE", "normal", "TRUE / FALSE."),
            ("referencia_comercializadora", "", "normal", ""),
            ("tension_normalizada", "", "normal", ""),
            ("tension_v", "", "normal", "Tensión en voltios (entero)."),
            ("modo_control_potencia", "icp", "normal", "icp / maximetro."),
            ("P1", "5.5", "normal", "Potencia P1 en kW. Hasta 15 kW: múltiplos de 0,1."),
            ("P2", "5.5", "normal", "Potencia P2 en kW."),
            ("P3", "", "normal", "Potencia P3 en kW."),
            ("P4", "", "normal", "Potencia P4 en kW."),
            ("P5", "", "normal", "Potencia P5 en kW."),
            ("P6", "", "normal", "Potencia P6 en kW."),
            ("agree_tarifa", "", "normal", "AAAA-MM-DD."),
            ("agree_dh", "", "normal", "AAAA-MM-DD."),
            ("agree_tensio", "", "normal", "AAAA-MM-DD."),
            ("agree_tipus", "", "normal", "AAAA-MM-DD."),
            ("es_autoconsumo", "FALSE", "normal", "TRUE / FALSE."),
            ("bono_social", "FALSE", "normal", "TRUE / FALSE."),
            ("vivienda_habitual", "", "normal", "TRUE / FALSE."),
            ("tipo_subseccion", "", "normal", ""),
            ("peaje_directo", "FALSE", "normal", "TRUE / FALSE."),
            ("telegestion", "FALSE", "normal", "TRUE / FALSE."),
            ("electrointensivo", "FALSE", "normal", "TRUE / FALSE."),
            ("codigo_solicitud_electrointensivo", "", "normal", ""),
            ("exencion_iese", "FALSE", "normal", "TRUE / FALSE."),
            ("no_cortable", "FALSE", "normal", "TRUE / FALSE."),
            ("no_cesion_sips", "FALSE", "normal", "TRUE / FALSE."),
            ("no_cesion_sips_fecha", "", "normal", "AAAA-MM-DD."),
            ("cie", "", "normal", "Certificado de Instalación Eléctrica."),
            ("cnae", "", "normal", "Código CNAE-2009."),
            ("notas", "", "normal", ""),
            ("activo", "TRUE", "normal", "TRUE / FALSE."),
        ],
    },
}

ORDEN_CARGA = ["titulares", "comercializadoras_empresa", "suministros", "contratos"]


def _hoja_instrucciones(wb: Workbook, spec: dict) -> None:
    ws = wb.active
    ws.title = "Instrucciones"
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 85

    t = ws.cell(row=1, column=1, value=spec["titulo"])
    t.font = Font(name=_FONT, bold=True, size=14)
    ws.merge_cells("A1:B1")

    def fila(r, a, b):
        ca = ws.cell(row=r, column=1, value=a)
        ca.font = Font(name=_FONT, bold=True, size=10)
        ca.alignment = Alignment(vertical="top", wrap_text=True)
        cb = ws.cell(row=r, column=2, value=b)
        cb.font = Font(name=_FONT, size=10)
        cb.alignment = Alignment(vertical="top", wrap_text=True)

    fila(3, "¿Qué es esto?", spec["que_es"])
    fila(4, "Orden de carga", spec["orden"])
    fila(5, "La fila 2", "Es un EJEMPLO. Bórrala o sobreescríbela. Tus datos van desde la fila 2 hacia abajo.")
    fila(6, "Colores", "Verde = clave única · Azul = enlace (debe existir ya) · Amarillo = automático (NO rellenar) · Gris = normal.")
    fila(7, "Si algo falla", "La migración carga lo correcto y devuelve un Excel con un resumen y el detalle de cada fila fallida (entidad, fila, columna y motivo). Corrige y vuelve a subir.")

    cab = ws.cell(row=9, column=1, value="COLUMNAS")
    cab.font = Font(name=_FONT, bold=True, size=11)
    r = 10
    for nombre, _ej, tipo, ayuda in spec["columnas"]:
        c = ws.cell(row=r, column=1, value=nombre)
        c.font = Font(name=_FONT, bold=True, size=10)
        c.fill = PatternFill("solid", fgColor=_FILL[tipo])
        c.alignment = Alignment(vertical="top", wrap_text=True)
        d = ws.cell(row=r, column=2, value=ayuda or "")
        d.font = Font(name=_FONT, size=10)
        d.alignment = Alignment(vertical="top", wrap_text=True)
        r += 1


def _hoja_datos(wb: Workbook, spec: dict) -> None:
    ws = wb.create_sheet(spec["hoja"])
    for i, (nombre, ejemplo, tipo, _ayuda) in enumerate(spec["columnas"], start=1):
        c = ws.cell(row=1, column=i, value=nombre)
        c.font = Font(name=_FONT, bold=True, size=10)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.fill = PatternFill("solid", fgColor=_FILL[tipo])
        if tipo in _COMMENT:
            c.comment = Comment(_COMMENT[tipo], "ERP")
        e = ws.cell(row=2, column=i, value=ejemplo)
        e.font = Font(name=_FONT, italic=True, size=10, color="808080")
        e.fill = PatternFill("solid", fgColor="F2F2F2")
        ws.column_dimensions[c.column_letter].width = max(12, len(nombre) + 2)
    ws.cell(row=1, column=1).comment = Comment(
        "Fila 2 = EJEMPLO. Tus datos desde la fila 2. Ver hoja 'Instrucciones'.", "ERP"
    )
    ws.freeze_panes = "A2"


def generar_plantilla(entidad: str) -> bytes:
    """Devuelve el .xlsx (bytes) de la plantilla de la entidad indicada."""
    spec = PLANTILLAS.get(entidad)
    if spec is None:
        raise ValueError(f"Entidad de plantilla desconocida: {entidad}")
    wb = Workbook()
    _hoja_instrucciones(wb, spec)
    _hoja_datos(wb, spec)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()
